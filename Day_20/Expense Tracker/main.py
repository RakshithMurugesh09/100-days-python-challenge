from openpyxl import Workbook, load_workbook
import re
from datetime import datetime
import os

FILE_NAME = "Expense Tracker.xlsx"

def add_expense():

    try:
        wb = load_workbook(FILE_NAME)
        ws = wb["Expense Tracker"]

    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Tracker"

        ws.append([
            "Date",
            "Category",
            "Description",
            "Amount"
        ])

    while True:
        user_date = input("Enter Date (DD-MM-YYYY): ")

        pattern = r"^(0[1-9]|[12][0-9]|3[01])-(0[1-9]|1[0-2])-(19\d{2}|20\d{2})$"

        if not re.match(pattern, user_date):
            print("Invalid Date Format")
            continue

        try:
            datetime.strptime(user_date, "%d-%m-%Y")
            break

        except ValueError:
            print("Invalid Date")

    while True:
        category = input("Enter Category: ").strip()

        if category:
            break

        print("Category cannot be empty.")

    while True:
        description = input("Enter Description: ").strip()

        if description:
            break

        print("Description cannot be empty.")

    while True:
        try:
            amount = float(input("Enter Amount: "))

            if amount <= 0:
                print("Amount must be greater than 0")
                continue

            break

        except ValueError:
            print("Invalid Amount")

    ws.append([user_date, category, description, amount])

    wb.save(FILE_NAME)

    print("✅ Expense Added Successfully")


def view_expenses():

    try:
        wb = load_workbook(FILE_NAME)
        ws = wb["Expense Tracker"]

        if ws.max_row == 1:
            print("No Expenses Found")
            return

        for row in ws.iter_rows(min_row=2, values_only=True):

            print(f"\nDate        : {row[0]}")
            print(f"Category    : {row[1]}")
            print(f"Description : {row[2]}")
            print(f"Amount      : ₹{row[3]}")

    except FileNotFoundError:
        print("Expense Tracker Excel File Not Created")


def show_total_spent():

    try:
        wb = load_workbook(FILE_NAME)
        ws = wb["Expense Tracker"]

        if ws.max_row == 1:
            print("No Expenses Found")
            return

        total = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            total += float(row[3])

        print(f"\n💰 Total Spent: ₹{total:.2f}")

    except FileNotFoundError:
        print("Expense Tracker Excel File Not Created")


while True:

    print("""
======== EXPENSE TRACKER ========

1. Add Expense
2. View Expenses
3. Show Total Spent
4. Exit

================================
""")

    choice = input("Enter Choice: ")

    if choice == "1":
        add_expense()

    elif choice == "2":
        view_expenses()

    elif choice == "3":
        show_total_spent()

    elif choice == "4":
        print("Thank You")
        break

    else:
        print("Invalid Choice")

print(os.path.abspath(FILE_NAME))